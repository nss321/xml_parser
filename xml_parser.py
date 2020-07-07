import xml.etree.ElementTree as Et
from lxml import etree
import openpyxl as xl
from os import listdir as dir
import re, os
import sys
#import html

# parser = etree.XMLParser(recover=True)
# parser = Et.XMLParser(encoding="utf-8")

save_path = 'C:\\Users\\user\\Downloads\\xml_test.xlsx'  # 엑셀파일 경로
# xml_path = 'C:\\Users\\user\\Downloads\\drive-download-20200623T064058Z-001\\'  # xml 파일 경로
xml_path = 'C:\\Users\\user\\Downloads\\XML 파일 업로드 (File responses)-20200706T071726Z-001\\XML 파일 업로드 (File responses)\\'

target = ['Operating System', 'CPU', 'RAM', 'Graphics', 'Storage']
Specific = ['Operating System', 'CPU', 'RAM', 'Graphics', 'Storage', 'Motherboard', 'Ex IP Address', 'NetBIOS Name']
result = []
tree_list = []
name_list = []

file_list = dir(xml_path)
file_list.sort()   # 파일명 불러옴

# print(file_list)

# def ignore_open(p):
#     temf = 'temp_file'
#     with open(temf, 'wt',encoding='utf-8') as temp:
#         o = open(p, 'r',encoding='utf-8')
#         temp.write(re.sub("&#(x?)([0-9a-fA-F]+);", '', o.read()))
#     rs = open(temf)
#     os.unlink(temf)
#     return rs

# test = "local&#x03;"
#
# print(html.unescape(test))
#
# testtree = Et.parse(testfile)
# print(testtree)

for i in file_list:
    name_list.append(i.split('.')[0])  # 엑셀 시트의 Name 행에 들어갈 파일명 분리

for i in file_list:
    x = xml_path + i  # xml 경로 + 파일이름으로 각각의 트리 객체를 만들어 리스트로 만듦
    print(html.unescape(x))
    # tree_tmp = Et.parse(x)
    tree_tmp = Et.parse(x)
    print(x)
    tree_list.append(tree_tmp)

print(tree_list)

for pxml in tree_list:
    for i in target:
        if i == 'CPU':
            data = pxml.find("./mainsection[@title='{}']/section".format(target[1]))
            if data.attrib['title'].find('AMD') != -1:
                data = pxml.find("./mainsection[@title='{}']/section/entry[@title='{}']".format(target[1],
                                                                                                'Specification'))
                result.append(data.attrib['value'])
            else:
                result.append(data.attrib['title'])
            continue
        elif i == 'Graphics':
            data = pxml.find("./mainsection[@title='Summary']/section[@title='{0}']/entry[2]".format(i))
        else:
            data = pxml.find("./mainsection[@title='Summary']/section[@title='{0}']/entry".format(i))

        result.append(data.attrib['title'])
    motherboard_manufacturer = pxml.find("./mainsection[@title='Motherboard']/entry[@title='Manufacturer']")
    motherboard_model = pxml.find("./mainsection[@title='Motherboard']/entry[@title='Model']")
    result.append(motherboard_manufacturer.attrib['value']+'_'+motherboard_model.attrib['value'])
    network_IP = pxml.find("./mainsection[@title='Network']/entry[@title='External IP Address']")
    network_BIOS = pxml.find("./mainsection[@title='Network']/section[@title='Computer Name']/"
                             "entry[@title='NetBIOS Name']")
    result.append(network_IP.attrib['value'])
    result.append(network_BIOS.attrib['value'])

# wb = xl.load_workbook()
wb = xl.load_workbook(str(save_path))
ws = wb['Sheet1']

NetBIOS_list =[]
for index, value in enumerate(result,start=1):
    if index % 8 == 0:
        NetBIOS_list.append(value)

ws_col = ws['j'][1:]
check_list = []
for i in ws_col:
    check_list.append(i.value)
cnt = 0
for i, j in zip(check_list, NetBIOS_list):
    if i == j:
        cnt += 1

if cnt > 0:
    sys.exit()

else:
    i = ws.max_row + 1
    print(len(result)//8)
    # Start = ['B2']    cell(1,1,'Name')~cell(1,9,'NetBIOS')
    for index, value in enumerate(result, start=1):
        if index % 8 == 0:
            ws.cell(row=i, column=10, value=value)
            i += 1
        else:
            ws.cell(row=i, column=(index % 8)+2, value=value)
    for i in name_list:
        ws.cell(row=name_list.index(i)+2,column=2,value=i)

wb.save('./xml_test.xlsx')
wb.close()
